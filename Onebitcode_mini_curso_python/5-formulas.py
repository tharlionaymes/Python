from openpyxl import load_workbook
from openpyxl.utils import get_column_letter # pega a letra da coluna

# 1 - Lê pasta de trabalho e planilha
wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatorio"]

# 2 - Ter referências das linhas e colunas, saber quantas linhas e colunas existe na planilha
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

""" 3 - Incluindo Fórmula 1 por 1
sheet ["B6"] = "=Sum(B2:B5)"
sheet ["B6"].style = "Currency"
"""

# 3 - Incluindo Fórmula automática

for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    teste = sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    teste2 = sheet[f"{letter}{max_row+1}"].style = "Currency"

sheet ["A6"] = "TOTAL" 
wb.save("test.xlsx")