from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference #gráfico de barra com as referencias

# 1 - Lê pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"]

# 2 - Ter referências das linhas e colunas, saber quantas linhas e colunas existe na planilha
min_column = wb.active.min_column
max_column = wb.active.max_column

min_row = wb.active.min_row
max_row = wb.active.max_row

# 3 - Adicionando Dados e Categorias no Gráfico
barchart = BarChart() # chamamos o barchart

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row,
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row,
)

barchart.add_data(data, titles_from_data = True)
barchart.set_categories(categories)

# 4 - Criando o Gráfico na planilha
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas Por Fabricantes"
barchart.style = 2

# 5 - Salvando o WorkBook (Pasta de Trabalho do Excel)

wb.save("data/barchart.xlsx")